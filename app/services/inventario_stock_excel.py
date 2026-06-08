# app/services/inventario_stock_excel.py
"""
Generador de planilla Excel para el módulo Inventario de Stock.

Separado del service principal para respetar SRP:
este archivo solo sabe cómo armar el Excel, no cómo obtener los datos.

Columnas de la planilla (1 fila por serie):
  Código | Descripción | Empresa | Serie | Lote | Vencimiento |
  Depósito | Estado sistema | En tránsito | Cant. Omnimedica |
  Cant. Finnegans | Presente físico (checkbox) | Observaciones

Retorna bytes listos para streamear desde FastAPI (StreamingResponse).
"""

import io
import logging
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from app.models.inventario_stock import InventarioRelevamiento, InventarioRelevamientoSerie

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Paleta de estilos
# ---------------------------------------------------------------------------
_COLOR_HEADER_BG = "1F4E79"   # azul oscuro corporativo
_COLOR_HEADER_FG = "FFFFFF"
_COLOR_TRANSITO_BG = "FFF2CC"  # amarillo suave para series en tránsito
_COLOR_VENCE_BG = "FCE4D6"     # naranja suave para lotes próximos a vencer
_COLOR_BORDE = "BFBFBF"

_THIN = Side(style="thin", color=_COLOR_BORDE)
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

# Columnas: (header_label, ancho_col)
_COLUMNAS = [
    ("Código", 14),
    ("Descripción", 30),
    ("Empresa", 20),
    ("Serie", 22),
    ("Lote", 16),
    ("Vencimiento", 14),
    ("Depósito", 18),
    ("Estado sistema", 15),
    ("En tránsito", 12),
    ("Cant. Omnimedica", 16),
    ("Cant. Finnegans", 15),
    ("Presente físico", 15),
    ("Observaciones", 30),
]


# ---------------------------------------------------------------------------
# Función pública
# ---------------------------------------------------------------------------


def generar_planilla_excel(
    relevamiento: InventarioRelevamiento,
    series: list[InventarioRelevamientoSerie],
) -> bytes:
    """
    Genera el Excel del relevamiento y retorna los bytes del archivo.

    Args:
        relevamiento: cabecera del ciclo (para el título).
        series: lista de series ya ordenadas.

    Returns:
        bytes del archivo .xlsx.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Relevamiento"

    _escribir_titulo(ws, relevamiento)
    _escribir_headers(ws)
    _escribir_filas(ws, series)
    _aplicar_anchos(ws)
    _freeze_panes(ws)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    logger.info(
        f"[Excel] Planilla generada: {len(series)} series | "
        f"relevamiento_id={relevamiento.id}"
    )
    return buffer.read()


# ---------------------------------------------------------------------------
# Helpers de construcción
# ---------------------------------------------------------------------------


def _escribir_titulo(ws, relevamiento: InventarioRelevamiento) -> None:
    """Fila 1: título del ciclo."""
    generado_en = datetime.now(timezone.utc).strftime("%d/%m/%Y %H:%M UTC")
    titulo = (
        f"Relevamiento de Stock — {relevamiento.proveedor} | "
        f"{relevamiento.mes_ciclo} | Generado: {generado_en}"
    )
    ws.merge_cells(f"A1:{get_column_letter(len(_COLUMNAS))}1")
    celda = ws["A1"]
    celda.value = titulo
    celda.font = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    celda.fill = PatternFill("solid", fgColor=_COLOR_HEADER_BG)
    celda.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 22


def _escribir_headers(ws) -> None:
    """Fila 2: encabezados de columnas."""
    header_font = Font(name="Calibri", bold=True, size=10, color=_COLOR_HEADER_FG)
    header_fill = PatternFill("solid", fgColor=_COLOR_HEADER_BG)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, (label, _) in enumerate(_COLUMNAS, start=1):
        celda = ws.cell(row=2, column=col_idx, value=label)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_align
        celda.border = _BORDER

    ws.row_dimensions[2].height = 30


def _escribir_filas(
    ws, series: list[InventarioRelevamientoSerie]
) -> None:
    """A partir de la fila 3: una fila por serie."""
    font_normal = Font(name="Calibri", size=10)
    align_center = Alignment(horizontal="center", vertical="center")
    align_left = Alignment(horizontal="left", vertical="center")

    for row_idx, serie in enumerate(series, start=3):
        en_transito = serie.en_transito
        row_fill = None
        if en_transito:
            row_fill = PatternFill("solid", fgColor=_COLOR_TRANSITO_BG)

        valores = [
            serie.codigo,
            serie.descripcion or "",
            serie.empresa or "",
            serie.serie,
            serie.lote or "",
            serie.vencimiento or "",
            serie.deposito or "",
            serie.estado_sistema,
            "SÍ" if en_transito else "NO",
            None,                          # Cant. Omnimedica — calculada post-análisis
            float(serie.cant_finnegans) if serie.cant_finnegans is not None else None,
            "",                            # Presente físico — checkbox manual
            serie.observaciones or "",
        ]

        for col_idx, valor in enumerate(valores, start=1):
            celda = ws.cell(row=row_idx, column=col_idx, value=valor)
            celda.font = font_normal
            celda.border = _BORDER
            # Alineación centrada para números y flags
            if col_idx in {9, 10, 11}:
                celda.alignment = align_center
            else:
                celda.alignment = align_left

            if row_fill:
                celda.fill = row_fill

        ws.row_dimensions[row_idx].height = 16


def _aplicar_anchos(ws) -> None:
    for col_idx, (_, ancho) in enumerate(_COLUMNAS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = ancho


def _freeze_panes(ws) -> None:
    """Congela título + headers para facilitar el scroll."""
    ws.freeze_panes = "A3"